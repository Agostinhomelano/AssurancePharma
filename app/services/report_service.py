from datetime import datetime, date
from io import BytesIO
from app import db
from app.models import Sale, SaleItem, Medicine, Employee, DailyReport

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ReportService:

    @staticmethod
    def generate_daily_excel(gerant_id, report_date=None):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl n'est pas installé. Installez-le avec: pip install openpyxl")

        if report_date is None:
            report_date = date.today()

        wb = openpyxl.Workbook()

        header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
        title_font = Font(name='Calibri', bold=True, size=14, color='0D6EFD')
        normal_font = Font(name='Calibri', size=11)
        bold_font = Font(name='Calibri', bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        sales = Sale.query.join(Employee).filter(
            db.func.date(Sale.created_at) == report_date,
            Employee.gerant_id == gerant_id
        ).order_by(Sale.created_at.desc()).all()

        medicines = Medicine.query.filter_by(gerant_id=gerant_id).all()
        employees = Employee.query.filter_by(gerant_id=gerant_id).all()

        total_sales = len(sales)
        total_amount = sum(s.total_amount for s in sales)
        total_profit = sum(s.profit for s in sales)
        low_stock = [m for m in medicines if m.stock_faible]
        expiring = [m for m in medicines if m.proche_expiration]

        # Sheet 1: Résumé
        ws_summary = wb.active
        ws_summary.title = "Résumé"
        ws_summary.merge_cells('A1:C1')
        ws_summary['A1'] = f"Rapport Journalier - {report_date.strftime('%d/%m/%Y')}"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].alignment = Alignment(horizontal='center')

        summary_data = [
            ('Total Ventes', total_sales),
            ('Montant Total (FCFA)', total_amount),
            ('Bénéfice Total (FCFA)', total_profit),
            ('Médicaments en Stock Faible', len(low_stock)),
            ('Médicaments proche Expiration', len(expiring)),
            ('Nombre d\'Employés', len(employees) if employees else 0),
        ]

        for i, (label, value) in enumerate(summary_data, start=3):
            ws_summary[f'A{i}'] = label
            ws_summary[f'A{i}'].font = bold_font
            ws_summary[f'B{i}'] = value
            ws_summary[f'B{i}'].font = normal_font
            ws_summary[f'A{i}'].border = thin_border
            ws_summary[f'B{i}'].border = thin_border

        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 20

        # Sheet 2: Ventes détaillées
        ws_sales = wb.create_sheet("Ventes")
        headers = ['ID', 'Employé', 'Montant (FCFA)', 'Coût', 'Bénéfice', 'Méthode', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws_sales.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row, sale in enumerate(sales, start=2):
            data = [
                sale.id,
                f"{sale.employee.prenom} {sale.employee.nom}" if sale.employee else 'N/A',
                sale.total_amount,
                sale.total_cost,
                sale.profit,
                sale.payment_method,
                sale.created_at.strftime('%d/%m/%Y %H:%M') if sale.created_at else ''
            ]
            for col, value in enumerate(data, 1):
                cell = ws_sales.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border
                if col in (3, 4, 5):
                    cell.number_format = '#,##0.00'

        ws_sales.column_dimensions['A'].width = 8
        ws_sales.column_dimensions['B'].width = 25
        ws_sales.column_dimensions['C'].width = 15
        ws_sales.column_dimensions['D'].width = 12
        ws_sales.column_dimensions['E'].width = 12
        ws_sales.column_dimensions['F'].width = 15
        ws_sales.column_dimensions['G'].width = 18

        # Sheet 3: Stock
        ws_stock = wb.create_sheet("Stock")
        headers2 = ['Médicament', 'Catégorie', 'Quantité', 'Stock Min', 'Prix Vente', 'Prix Achat', 'Stock Faible', 'Expire bientôt']
        for col, header in enumerate(headers2, 1):
            cell = ws_stock.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row, med in enumerate(medicines, start=2):
            data = [
                med.nom, med.categorie, med.quantite, med.stock_minimum,
                med.prix_vente, med.prix_achat,
                'Oui' if med.stock_faible else 'Non',
                'Oui' if med.proche_expiration else 'Non'
            ]
            for col, value in enumerate(data, 1):
                cell = ws_stock.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border
                if med.stock_faible and col in (7,):
                    cell.font = Font(name='Calibri', size=11, color='DC3545', bold=True)

        ws_stock.column_dimensions['A'].width = 25
        ws_stock.column_dimensions['B'].width = 18
        ws_stock.column_dimensions['C'].width = 10
        ws_stock.column_dimensions['D'].width = 10
        ws_stock.column_dimensions['E'].width = 12
        ws_stock.column_dimensions['F'].width = 12
        ws_stock.column_dimensions['G'].width = 12
        ws_stock.column_dimensions['H'].width = 14

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        report = DailyReport(
            date=report_date,
            total_sales=total_sales,
            total_amount=total_amount,
            total_profit=total_profit,
            medicines_out_of_stock=len(low_stock),
            medicines_expiring=len(expiring),
            gerant_id=gerant_id
        )
        db.session.add(report)
        db.session.commit()

        return output, f"rapport_journalier_{report_date.strftime('%Y%m%d')}.xlsx"

    @staticmethod
    def generate_period_excel(gerant_id, start_date, end_date):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl n'est pas installé")

        wb = openpyxl.Workbook()
        header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
        title_font = Font(name='Calibri', bold=True, size=14, color='0D6EFD')
        normal_font = Font(name='Calibri', size=11)
        bold_font = Font(name='Calibri', bold=True, size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        sales = Sale.query.join(Employee).filter(
            Sale.created_at >= datetime.combine(start_date, datetime.min.time()),
            Sale.created_at <= datetime.combine(end_date, datetime.max.time()),
            Employee.gerant_id == gerant_id
        ).order_by(Sale.created_at.desc()).all()

        ws = wb.active
        ws.title = f"Rapport {start_date.strftime('%d/%m')}-{end_date.strftime('%d/%m')}"
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Rapport de Ventes du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        total_amount = sum(s.total_amount for s in sales)
        total_profit = sum(s.profit for s in sales)
        summary = [
            ('Nombre de ventes', len(sales)),
            ('Montant total (FCFA)', total_amount),
            ('Bénéfice total (FCFA)', total_profit),
            ('Moyenne par vente (FCFA)', total_amount / len(sales) if sales else 0),
        ]
        for i, (label, value) in enumerate(summary, start=3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = bold_font
            ws[f'B{i}'] = value
            ws[f'B{i}'].font = normal_font
            ws[f'A{i}'].border = thin_border
            ws[f'B{i}'].border = thin_border

        start_row = 8
        headers = ['Date', 'Employé', 'Articles', 'Montant', 'Bénéfice', 'Méthode']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row, sale in enumerate(sales, start=start_row + 1):
            data = [
                sale.created_at.strftime('%d/%m/%Y') if sale.created_at else '',
                f"{sale.employee.prenom} {sale.employee.nom}" if sale.employee else 'N/A',
                len(sale.items),
                sale.total_amount,
                sale.profit,
                sale.payment_method
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"rapport_ventes_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
